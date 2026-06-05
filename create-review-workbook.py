import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.datavalidation import DataValidation

ROOT = Path(__file__).resolve().parent
rows = json.loads((ROOT / "needs-review-items.json").read_text(encoding="utf-8"))

wb = Workbook()
ws = wb.active
ws.title = "Needs Review"

headers = [
    "school",
    "major",
    "requirementType",
    "reason",
    "rawText",
    "suggestedAction",
    "canonicalCourseId",
    "canonicalCourseName",
    "conditionType",
    "reviewerNote",
]
ws.append(headers)
for row in rows:
    ws.append([row.get(header, "") for header in headers])

header_fill = PatternFill("solid", fgColor="0F2F44")
header_font = Font(color="FFFFFF", bold=True)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(vertical="center")

widths = {
    "A": 24,
    "B": 34,
    "C": 15,
    "D": 34,
    "E": 72,
    "F": 24,
    "G": 24,
    "H": 30,
    "I": 24,
    "J": 42,
}
for col, width in widths.items():
    ws.column_dimensions[col].width = width
for row in ws.iter_rows(min_row=2):
    row[4].alignment = Alignment(wrap_text=True, vertical="top")
    row[9].alignment = Alignment(wrap_text=True, vertical="top")
ws.freeze_panes = "A2"
ws.auto_filter.ref = ws.dimensions

action_validation = DataValidation(
    type="list",
    formula1='"map_to_common_course,choice_condition,later_requirement,ignore_note,add_new_common_course,keep_review"',
    allow_blank=True,
)
condition_validation = DataValidation(
    type="list",
    formula1='"required,recommended,choice_any_one,choice_pick_n,enrollment_or_later,ignore"',
    allow_blank=True,
)
ws.add_data_validation(action_validation)
ws.add_data_validation(condition_validation)
last_row = max(2, len(rows) + 1)
action_validation.add(f"F2:F{last_row}")
condition_validation.add(f"I2:I{last_row}")

guide = wb.create_sheet("How To Review")
guide_rows = [
    ["Column", "What to enter"],
    ["suggestedAction", "Choose how this raw requirement should be handled."],
    ["canonicalCourseId", "If it maps to an existing common course, enter the course id."],
    ["canonicalCourseName", "Human-readable common course name."],
    ["conditionType", "Use choice_any_one for OR/pick-one rules; enrollment_or_later for not-before-application requirements."],
    ["reviewerNote", "Any clarification, new course suggestion, or data issue."],
]
for row in guide_rows:
    guide.append(row)
for cell in guide[1]:
    cell.fill = header_fill
    cell.font = header_font
guide.column_dimensions["A"].width = 24
guide.column_dimensions["B"].width = 100

actions = wb.create_sheet("Allowed Values")
actions.append(["suggestedAction", "Meaning"])
for row in [
    ["map_to_common_course", "This raw text should count as one existing common course."],
    ["choice_condition", "This is an OR / pick-one / choose-one rule."],
    ["later_requirement", "This is not a pre-application requirement; show separately."],
    ["ignore_note", "This is a note, source, policy sentence, or non-course item."],
    ["add_new_common_course", "Create a new common course in course-catalog.js."],
    ["keep_review", "Still ambiguous after review."],
]:
    actions.append(row)
for cell in actions[1]:
    cell.fill = header_fill
    cell.font = header_font
actions.column_dimensions["A"].width = 28
actions.column_dimensions["B"].width = 82

out = ROOT / "needs-review-items.xlsx"
wb.save(out)
print(out)
